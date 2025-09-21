import os
import csv
import shutil
import argparse
from datetime import datetime
import openpyxl

class PomeloStatusUpdater:
    CONFIG_COLUMNS = {
        "STATUS": 0,
        "FOLDER_PATH": 2,
        "INCLUDE": 4
    }

    def __init__(self, csv_path, backups_path, excel_path):
        self.csv_path = csv_path
        self.backups_path = backups_path
        self.excel_path = excel_path
        self.status_folders = self.load_status_folders_from_config()
        self.class_stats = {cls: {"rewrites": 0, "deletions": 0} for cls in self.status_folders.keys()}
        self.not_found_images = []
        self.rows = []
        self.header = []
        self.csv_lookup = {}
        self.image_class_map = {}

    def load_status_folders_from_config(self):
        status_folders = {}
        with open("configs\\image_statuses.csv", 'r', newline='') as config_file:
            reader = csv.reader(config_file)
            next(reader)
            for row in reader:
                if (len(row) > self.CONFIG_COLUMNS["INCLUDE"] and 
                    row[self.CONFIG_COLUMNS["INCLUDE"]].upper() == "TRUE"):
                    if (len(row) > self.CONFIG_COLUMNS["FOLDER_PATH"] and 
                        row[self.CONFIG_COLUMNS["FOLDER_PATH"]]):
                        status_name = row[self.CONFIG_COLUMNS["STATUS"]]
                        folder_path = row[self.CONFIG_COLUMNS["FOLDER_PATH"]]
                        status_folders[status_name] = folder_path
        return status_folders

    def make_backup(self):
        file_name = os.path.split(self.csv_path)[1]
        csv_name, extension = os.path.splitext(file_name)
        now = datetime.now()
        timestamp = now.strftime("%Y%m%d%H%M%S")
        csv_name = f"{csv_name}_backup_{timestamp}"
        new_path = os.path.join(self.backups_path, f"{csv_name}{extension}")
        os.makedirs(self.backups_path, exist_ok=True)
        shutil.copy2(self.csv_path, new_path)
        print(f"Backup created at: {new_path}")
        return new_path

    def load_csv(self):
        with open(self.csv_path, newline="", encoding="utf-8") as f:
            reader = list(csv.reader(f))
            self.header = reader[0]
            self.rows = reader[1:]

        self.csv_lookup = {row[0]: i for i, row in enumerate(self.rows)}

        while len(self.header) < 3:
            self.header.append(f"col{len(self.header)+1}")

    def process_status_folders(self):
        original_class_map = {}
        for class_name, folder in self.status_folders.items():
            if not os.path.isdir(folder):
                print(f"Warning: folder '{folder}' not found. Skipping...")
                continue
            for file in os.listdir(folder):
                if not file.lower().endswith(('.jpg', '.jpeg', '.png', '.bmp', '.tiff', '.tif')):
                    continue
                file_path = os.path.join(folder, file)
                if not os.path.isfile(file_path):
                    continue
                image_name, _ = os.path.splitext(file)
                if image_name in self.csv_lookup:
                    row_index = self.csv_lookup[image_name]
                    row = self.rows[row_index]
                    while len(row) < 3:
                        row.append("")
                    current_class_in_csv = row[2]
                    if image_name not in original_class_map and current_class_in_csv and current_class_in_csv != class_name:
                        original_class_map[image_name] = current_class_in_csv
                    if image_name in self.image_class_map:
                        prev_class = self.image_class_map[image_name]
                        if prev_class != class_name:
                            prev_folder = self.status_folders[prev_class]
                            prev_file_path = os.path.join(prev_folder, file)
                            if os.path.exists(prev_file_path):
                                os.remove(prev_file_path)
                                self.class_stats[prev_class]["deletions"] += 1
                                self.class_stats[prev_class]["rewrites"] -= 1
                    original_class = original_class_map.get(image_name, current_class_in_csv)
                    if class_name != original_class:
                        row[2] = class_name
                        self.class_stats[class_name]["rewrites"] += 1
                        self.image_class_map[image_name] = class_name
                    else:
                        self.image_class_map[image_name] = class_name
                else:
                    self.not_found_images.append(image_name)

    def save_csv(self):
        with open(self.csv_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(self.header)
            writer.writerows(self.rows)

    def export_to_excel(self):
        if os.path.exists(self.excel_path):
            wb = openpyxl.load_workbook(self.excel_path)
            ws = wb.worksheets[0]
            ws.delete_rows(1, ws.max_row)
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
        ws.append(self.header)
        for row in self.rows:
            ws.append(row)
        wb.save(self.excel_path)
        print(f"Excel file saved at: {self.excel_path}")

    def print_results(self):
        print("\n=== Rewrite & Deletion Summary ===")
        for cls, stats in self.class_stats.items():
            print(f"{cls}: {stats['rewrites']} rewrites, {stats['deletions']} deletions")

        if self.not_found_images:
            print("\n=== Images not found in CSV ===")
            for img in self.not_found_images:
                print(img)
        else:
            print("\nNo missing images.")

    def run(self):
        print(f"Loaded {len(self.status_folders)} classes from configuration:")
        for cls, folder in self.status_folders.items():
            print(f"  {cls}: {folder}")
        self.make_backup()
        self.load_csv()
        self.process_status_folders()
        self.save_csv()
        self.export_to_excel()
        self.print_results()


def run_pomelo_status_updater(csv_path, backups_path, excel_path):
    try:
        updater = PomeloStatusUpdater(csv_path, backups_path, excel_path)
        updater.run()
    except FileNotFoundError as e:
        print(f"Error: {e}")
        print("Please make sure the configuration file exists at the specified path")
    except Exception as e:
        print(f"An error occurred: {e}")


def main():
    parser = argparse.ArgumentParser(description='Update image statuses in CSV based on folder organization')
    parser.add_argument('--csv_path', default=r"tracker\tracker.csv", 
                       help='Path to the main CSV file (default: tracker\\tracker.csv)')
    parser.add_argument('--backups_path', default=r"tracker\backups", 
                       help='Path to backup directory (default: tracker\\backups)')
    parser.add_argument('--excel_path', default=r"tracker\stats.xlsx", 
                       help='Path to Excel output file (default: tracker\\stats.xlsx)')
    args = parser.parse_args()
    run_pomelo_status_updater(args.csv_path, args.backups_path, args.excel_path)


if __name__ == "__main__":
    main()
